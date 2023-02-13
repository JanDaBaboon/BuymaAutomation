from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import keyboard
from openpyxl import workbook, load_workbook
import os 
from os import listdir


#-----------------------------------------------------------------------------
#temp code
def pause():
    while True:
        if keyboard.read_key() == 'space':
            break
        
        
#-----------------------------------------------------------------------------
#information for website, driver, and xlsx
url = "https://www.buyma.com/login/"
url2 = "https://www.buyma.com/my/sell/new?tab=b"
driver = webdriver.Chrome(executable_path="./chromedriver")
wb = load_workbook(r"C:\Users\junya\Desktop\BuymaAutomation\xlsx\inventory_data.xlsx")
ws = wb["inventory_data_sheet"]
image_location = r"C:\Users\junya\Desktop\BuymaAutomation\images"

#selection elements in order of uses for different input elements
login_id = "txtLoginId"
password_id = "txtLoginPass"
submit_element_id = "login_do"
skip_button_class = "driver-close-btn"
drag_image_xpath = "/html/body/div[3]/div[2]/div[1]/div/div[1]/div/div/div/div[2]/form/div[1]/div/div/div[2]/div/div/div[1]/div/div/div/input"
drag_image_xpath2 = "/html/body/div[3]/div[2]/div[1]/div/div[1]/div/div/div/div[2]/form/div[1]/div/div/div[2]/div/div/div[1]/div/div/div[2]/input"
item_name_xpath = "/html/body/div[3]/div[2]/div[1]/div/div[1]/div/div/div/div[2]/form/div[2]/div[1]/div/div[2]/div/div/div[1]/input"
item_comment_xpath = "/html/body/div[3]/div[2]/div[1]/div/div[1]/div/div/div/div[2]/form/div[2]/div[2]/div/div[2]/div/div/div[1]/textarea"

#lists used for content management
item_name_list = []
item_comment_list = []
item_category1_list = []
item_category2_list = []
item_category3_list = []

#-----------------------------------------------------------------------------
#Functions for navigation
def Login(input, id):
    field = driver.find_element(By.ID, id)
    for x in input :
        field.send_keys(x)
    return driver

#Clicking function that clicks on specified ID
def Click_ID(id) :
    driver.find_element(By.ID, id).click()
    return driver

#Clicking function that clicks on specified Class
def Click_Class(class_id) : 
    driver.find_element(By.CLASS_NAME, class_id).click()
    return driver

#Clicking fucntion that clicks on specified Xpath
def Click_Xpath(xpath) : 
    driver.find_element(By.XPATH, xpath).click()
    return driver
    
#Function to upload photos. Iliterates through each image in given item's folder. Changes Xpath input according to current photo count. 
def Upload_Photo_xPath(xPath) :
    
    current_item_image_location = os.path.join(image_location, item_name_list[item_pointer])

    photo_counter = 1
    
    for images in os.listdir(current_item_image_location):
        
        i = os.path.join(current_item_image_location, images)
        
        if photo_counter == 1:
            driver.find_element(By.XPATH, xPath).send_keys(i)
            photo_counter =+ 1
            time.sleep(2)
            
        else :
            xPath2 = ("/html/body/div[3]/div[2]/div[1]/div/div[1]/div/div/div/div[2]/form/div[1]/div/div/div[2]/div/div/div[1]/div/div/div[%s]/input" % photo_counter)
            driver.find_element(By.XPATH, xPath2).send_keys(i)
            time.sleep(1)
            return driver
        

#Function to access xlsx sheet. Returns a list with all the items of first row.
def AccessSheet(sheet) :
    for col in sheet.iter_cols(min_row=2, min_col=2, max_col=2, max_row=100):
        for cell in col:
            if cell.value != None :
                item_name_list.append(cell.value)
                print(item_name_list)    
    
    for col in sheet.iter_cols(min_row=2, min_col=3, max_col=3, max_row=100):
        for cell in col:
            if cell.value != None :
                item_comment_list.append(cell.value)
                print(item_comment_list)         
                
    for col in sheet.iter_cols(min_row=2, min_col=4, max_col=4, max_row=100):
        for cell in col:
            if cell.value != None :
                item_category1_list.append(cell.value)
                print(item_category1_list)
            elif cell.value == None :
                item_category1_list.append(0)
                print(item_category1_list)
                
    for col in sheet.iter_cols(min_row=2, min_col=5, max_col=5, max_row=100):
        for cell in col:
            if cell.value != None :
                item_category2_list.append(cell.value)
                print(item_category2_list)
            elif cell.value == None :
                item_category1_list.append(0)
                print(item_category1_list)
                
    for col in sheet.iter_cols(min_row=2, min_col=6, max_col=6, max_row=100):
        for cell in col:
            if cell.value != None :
                item_category3_list.append(cell.value)
                print(item_category3_list)
            elif cell.value == None :
                item_category1_list.append(0)
                print(item_category1_list)
                
                                       
                
#Function to type in item name                
def TypeItemName(xpath) :
   driver.find_element(By.XPATH, xpath).send_keys(item_name_list[item_pointer])
   
#Function to type in item comment              
def TypeItemComment(xpath) :
   driver.find_element(By.XPATH, xpath).send_keys(item_comment_list[item_pointer])  
   

def NavigateCategory(cat1, cat2, cat3) :
    return
    
   
   
   
   
         
# -----------------------------------------------------------------------------
#Tracking which item we are on. Change the counter to change item.
item_pointer = 0

# -----------------------------------------------------------------------------
# Order of operations

driver.get(url)    
driver.maximize_window()

Login("kokomimi.buyma@gmail.com", login_id)

Login("koko12345", password_id)

Click_ID(submit_element_id)

driver.get(url2)

Click_Class(skip_button_class)

AccessSheet(ws)

Upload_Photo_xPath(drag_image_xpath)

TypeItemName(item_name_xpath)

TypeItemComment(item_comment_xpath)


pause()
time.sleep(1)




